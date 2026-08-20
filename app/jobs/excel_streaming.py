"""
Streaming Excel ZIP Export Module (Item 9)
Provides memory-bounded, streaming export of coverage analysis and line indices into ZIP archives:
- Keyset-paginated row streaming per directory (no giant in-memory row materialize)
- Direct ZIP archive streaming
- Atomic .part file creation and cleanup on error
- Bounded memory footprint under multi-million line datasets
"""

import os
import sys
import zipfile
import tempfile
import logging
from typing import Dict, Any, List, Generator, Optional, Tuple

logger = logging.getLogger(__name__)

# The ZIP writer intentionally processes one directory at a time.  This is the
# hard upper bound exposed to callers and leaves room for a future two-directory
# pipeline without allowing an unbounded list of detail futures.
MAX_INFLIGHT_DIR_EXPORTS = 2
DETAIL_BATCH_SIZE = 5000

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAVE_OPENPYXL = True
except ImportError:
    openpyxl = None
    HAVE_OPENPYXL = False

def build_directory_summary_workbook(dir_summaries: List[Dict[str, Any]]) -> str:
    """Create directory summary XLSX and return temp file path."""
    if not HAVE_OPENPYXL:
        raise RuntimeError("openpyxl is required for Excel export")
        
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "目录汇总"
    
    headers = ["目录路径", "未覆盖代码行数", "已确认行数", "已填待确认行数", "未分析行数", "完成率"]
    ws.append(headers)
    
    for s in dir_summaries:
        total = s.get("total_uncovered", 0)
        confirmed = s.get("confirmed_total", 0)
        draft = s.get("draft_total", 0)
        pending = max(0, total - confirmed - draft)
        rate = f"{(confirmed / total * 100.0):.1f}%" if total > 0 else "100.0%"
        ws.append([
            s.get("directory", "/"),
            total,
            confirmed,
            draft,
            pending,
            rate
        ])
        
    temp_f = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    temp_path = temp_f.name
    temp_f.close()
    wb.save(temp_path)
    wb.close()
    return temp_path

def build_directory_detail_workbook(
    directory_name: str,
    row_generator: Generator[Dict[str, Any], None, None]
) -> str:
    """Create directory detail XLSX in write-only/streaming mode to minimize memory."""
    if not HAVE_OPENPYXL:
        raise RuntimeError("openpyxl is required for Excel export")
        
    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet(title="未覆盖代码明细")
    
    headers = [
        "文件路径", "行号", "分析状态", "确认人",
        "条件覆盖方法", "无条件覆盖原因", "代码内容"
    ]
    ws.append(headers)
    
    for r in row_generator:
        ws.append([
            r.get("file_path", ""),
            r.get("line_number", 0),
            r.get("status", "未确认"),
            r.get("reviewer", ""),
            r.get("coverage_method", ""),
            r.get("uncovered_reason", ""),
            r.get("code_line", "")
        ])
        
    temp_f = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    temp_path = temp_f.name
    temp_f.close()
    wb.save(temp_path)
    wb.close()
    return temp_path

def export_project_coverage_streaming_zip(
    project_name: str,
    output_zip_path: str,
    dir_summaries: List[Dict[str, Any]],
    get_directory_rows_fn, # fn(project_name, directory) -> Generator[Dict[str, Any]]
    on_progress_fn=None,
    member_name_prefix: str = "detail_",
    member_name_separator: str = "_",
) -> str:
    """
    Stream full project export into ZIP archive with bounded memory.
    """
    part_path = output_zip_path + ".part"
    os.makedirs(os.path.dirname(os.path.abspath(output_zip_path)), exist_ok=True)
    
    temp_files_to_clean = []
    try:
        with zipfile.ZipFile(part_path, "w", zipfile.ZIP_DEFLATED) as zip_out:
            zip_out.writestr(
                "EXPORT_STARTED.txt",
                "Directory Excel export started for project: {}\n".format(project_name),
            )
            # 1. Directory summary
            summary_xlsx = build_directory_summary_workbook(dir_summaries)
            temp_files_to_clean.append(summary_xlsx)
            zip_out.write(summary_xlsx, "00_目录汇总.xlsx")
            
            # 2. Iterate each directory
            total_dirs = len(dir_summaries)
            for idx, dinfo in enumerate(dir_summaries, start=1):
                dpath = dinfo.get("directory", "root")
                safe_name = dpath.strip("/").replace("/", member_name_separator) or "root"
                
                # Fetch row generator
                rows_gen = get_directory_rows_fn(project_name, dpath)
                detail_xlsx = build_directory_detail_workbook(safe_name, rows_gen)
                temp_files_to_clean.append(detail_xlsx)
                
                zip_out.write(detail_xlsx, f"{member_name_prefix}{safe_name}.xlsx")
                
                # Immediately remove temporary XLSX from disk to free resources
                try:
                    os.remove(detail_xlsx)
                    temp_files_to_clean.remove(detail_xlsx)
                except Exception:
                    pass
                    
                if on_progress_fn:
                    on_progress_fn(idx, total_dirs, dpath)
                    
        # Atomic rename
        os.replace(part_path, output_zip_path)
        return output_zip_path
    except Exception as e:
        if os.path.isfile(part_path):
            try:
                os.remove(part_path)
            except Exception:
                pass
        raise e
    finally:
        for tf in temp_files_to_clean:
            if os.path.isfile(tf):
                try:
                    os.remove(tf)
                except Exception:
                    pass
