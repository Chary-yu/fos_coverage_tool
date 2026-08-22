"""
Targeted Tests for Phase 3 Background Jobs & Export Resource Governance (Items 6, 9)
"""

import unittest
import os
import sys
import time
import tempfile
import shutil
import zipfile
from unittest import mock

_REPO_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
if _REPO_ROOT not in sys.path:
    sys.path.insert(0, _REPO_ROOT)

from app.jobs.bounded_executor import BoundedJobExecutor, STATUS_COMPLETED, STATUS_CANCELLED
from app.jobs.excel_streaming import (
    build_directory_detail_workbook,
    export_project_coverage_streaming_zip,
)
from app.services import export_service as export_service_module
from app.services.export_service import ExportService

class TestPhase3JobsExport(unittest.TestCase):

    def setUp(self):
        self.test_dir = tempfile.mkdtemp()

    def tearDown(self):
        shutil.rmtree(self.test_dir, ignore_errors=True)

    def test_item_6_bounded_job_executor(self):
        """Verify BoundedJobExecutor execution, duplicate reuse, and cancellation."""
        executor = BoundedJobExecutor(max_workers=2, max_queue_size=10)
        
        def slow_task(duration):
            time.sleep(duration)
            return "done"
            
        # Submit task 1
        job1 = executor.submit_job("test_task", "ProjA", slow_task, args=(0.1,))
        # Submit duplicate task -> reuses job1
        job2 = executor.submit_job("test_task", "ProjA", slow_task, args=(0.1,))
        self.assertEqual(job1.job_id, job2.job_id)
        
        # Wait for completion
        time.sleep(0.3)
        self.assertEqual(job1.status, STATUS_COMPLETED)
        self.assertEqual(job1.result, "done")
        
        # Test cancellation
        job_cancel = executor.submit_job("long_task", "ProjB", slow_task, args=(1.0,), reuse_existing=False)
        ok = executor.cancel_job(job_cancel.job_id)
        self.assertTrue(ok)
        self.assertTrue(job_cancel.cancel_requested)
        
        executor.shutdown(wait=False)

    def test_item_9_excel_zip_streaming_export(self):
        """Verify streaming export creates valid multi-sheet zip archive with summary & detail."""
        out_zip = os.path.join(self.test_dir, "export_test.zip")
        
        dir_summaries = [
            {"directory": "src/core", "total_uncovered": 50, "confirmed_total": 30, "draft_total": 5},
            {"directory": "src/net", "total_uncovered": 20, "confirmed_total": 10, "draft_total": 0}
        ]
        
        def mock_row_generator(project, directory):
            count = 50 if directory == "src/core" else 20
            for i in range(1, count + 1):
                yield {
                    "file_path": f"{directory}/file.c",
                    "line_number": i,
                    "status": "可覆盖" if i <= 10 else "未确认",
                    "reviewer": "Bob" if i <= 10 else "",
                    "coverage_method": "Unit test" if i <= 10 else "",
                    "uncovered_reason": "",
                    "code_line": f"int x_{i} = {i};"
                }
                
        zip_res = export_project_coverage_streaming_zip(
            project_name="TestProj",
            output_zip_path=out_zip,
            dir_summaries=dir_summaries,
            get_directory_rows_fn=mock_row_generator
        )
        
        self.assertTrue(os.path.isfile(out_zip))
        self.assertFalse(os.path.isfile(out_zip + ".part"))
        
        # Inspect contents of generated ZIP
        with zipfile.ZipFile(out_zip, "r") as z:
            names = z.namelist()
            self.assertIn("00_目录汇总.xlsx", names)
            self.assertIn("detail_src_core.xlsx", names)
            self.assertIn("detail_src_net.xlsx", names)

    def test_excel_export_keeps_confirmed_and_suggested_reviewers_separate(self):
        path = build_directory_detail_workbook(
            "src/core",
            iter([{
                "file_path": "src/core/a.c", "line_number": 1,
                "status": "可覆盖", "reviewer": "db-carol",
                "suggested_reviewer": "git-alice", "coverage_method": "unit",
                "uncovered_reason": "", "code_line": "return 0;",
            }]),
        )
        try:
            import openpyxl
            workbook = openpyxl.load_workbook(path, read_only=True)
            row = next(workbook.active.iter_rows(values_only=True))
            self.assertEqual(row[3], "确认 Reviewer")
            self.assertEqual(row[4], "Suggested Reviewer")
            values = next(workbook.active.iter_rows(min_row=2, values_only=True))
            self.assertEqual(values[3], "db-carol")
            self.assertEqual(values[4], "git-alice")
            workbook.close()
        finally:
            if os.path.isfile(path):
                os.remove(path)

    def test_jsonl_temp_is_removed_when_archive_write_fails(self):
        """The JSONL spill file is cleaned even after archive assembly fails."""
        output_path = os.path.join(self.test_dir, "vnext-export.zip")
        project_repo = mock.Mock()
        project_repo.get_project_by_name.return_value = {"id": 7}
        project_repo.get_scan.return_value = {"id": 11, "project_id": 7}
        project_repo.get_report_for_scan.return_value = {"report_id": "report-11"}
        project_repo.list_repository_snapshots.return_value = []
        project_repo.iter_scan_export_rows.return_value = iter([{
            "file_path": "src/main.c", "line_number": 1,
            "status": "未确认", "reviewer": "", "suggested_reviewer": "",
        }])
        service = ExportService(project_repo, self.test_dir)
        created_paths = []
        real_named_temporary_file = export_service_module.tempfile.NamedTemporaryFile

        def recording_named_temporary_file(*args, **kwargs):
            handle = real_named_temporary_file(*args, **kwargs)
            created_paths.append(handle.name)
            return handle

        with mock.patch.object(
            export_service_module.tempfile,
            "NamedTemporaryFile",
            side_effect=recording_named_temporary_file,
        ), mock.patch.object(
            export_service_module.zipfile.ZipFile,
            "write",
            side_effect=OSError("archive write failed"),
        ):
            with self.assertRaisesRegex(OSError, "archive write failed"):
                service.export_scan(object(), "fixture", 11,
                                    report_id="report-11", output_path=output_path)

        self.assertEqual(len(created_paths), 1)
        self.assertFalse(os.path.exists(created_paths[0]))
        self.assertFalse(os.path.exists(output_path + ".part"))

if __name__ == "__main__":
    unittest.main()
